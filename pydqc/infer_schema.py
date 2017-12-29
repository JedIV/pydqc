import pandas as pd
import numpy as np
import os

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Protection, NamedStyle, Alignment
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

from sklearn.externals.joblib import Parallel, delayed
import xlsxwriter


"""
function: infer data type based on dropna sample values
parameters:
sample_data: numpy array
	sample values
col: string
	column name
type_threshold: float
	threshold for assigning data type
"""
def _infer_dtype(sample_data, col, type_threshold):
	# get basic dtype from pandas
	col_dtype = str(pd.Series(sample_data).dtype)

	try:
		# date or str
		if (col_dtype == 'object') or (col_dtype == 'bool'):
			date_sample = pd.to_datetime(sample_data, errors='coerce')
			date_nan_per = np.sum(pd.isnull(date_sample)) * 1.0 / len(date_sample)
			if date_nan_per < (1.0 - type_threshold):
				sample_type = 'date'
			else:
				sample_type = 'str'
		elif 'datetime' in col_dtype:
			sample_type = 'date'
		else:
			sample_type = 'numeric'
	except:
		sample_type = 'error'

	type_info = {
		'column': col,
		'type': sample_type
	}
	return type_info


"""
function: calculate simple statistic information for each column
parameters:
sample_data: numpy array
	sample values
col: string
	column name
col_type: string
	column data type
"""
def _cal_column_stat(sample_data, col, col_type):
	col_stat = {}
	col_stat['column'] = col

	# get sample value of the column from not nan values
	if len(sample_data) > 5:
		col_stat['sample_value'] = str(list(np.random.choice(sample_data, 5)))
	elif len(sample_data) > 0:
		col_stat['sample_value'] = str(list(sample_data))
	else:
		col_stat['sample_value'] = np.nan
		
	# for numeric column
	# sample_nan_per, sample_num_uni, sample_min, sample_median, sample_max, sample_std
	if col_type == 'numeric':
		sample_data = pd.to_numeric(sample_data, errors='coerce')

		if len(sample_data) > 0:
			col_stat['sample_num_uni'] = len(np.unique(sample_data))
			col_stat['sample_min'] = np.min(sample_data)
			col_stat['sample_median'] = np.median(sample_data)
			col_stat['sample_max'] = np.max(sample_data)
			col_stat['sample_std'] = np.std(sample_data)
		else:
			col_stat['sample_num_uni'] = 0
			col_stat['sample_min'] = np.nan
			col_stat['sample_median'] = np.nan
			col_stat['sample_max'] = np.nan
			col_stat['sample_std'] = np.nan

	# for date and str column
	# only sample_nan_per and sample_num_uni
	else:
		if len(sample_data) > 0:
			col_stat['sample_num_uni'] = len(np.unique(sample_data))
		else:
			col_stat['sample_num_uni'] = 0
	return col_stat


"""
function: adjust column width and font family for sheet
parameters:
ws: excel worksheet
"""
def _adjust_column(ws):
	col_widths = {}
	for i, col in enumerate(ws.columns):
		col_name = xlsxwriter.utility.xl_col_to_name(i)
		col_widths[col_name] = 0
		for cell in col:
			cell.alignment = Alignment(horizontal='left', wrap_text=True)
			if cell:
				try:
					cell_length = len(str(cell.value))
				except:
					cell_length = len(cell.value)
				if cell_length > col_widths[col_name]:
					col_widths[col_name] = cell_length

	for key in col_widths.keys():
		col_widths[key] *= 1.5

	for i, col in enumerate(range(ws.max_column)):
		col_name = xlsxwriter.utility.xl_col_to_name(i)
		ws.column_dimensions[col_name].width = np.min([col_widths[col_name], 100])

	for col in ws.iter_cols(max_col=ws.max_column, min_row=ws.max_row, max_row=ws.max_row):
		for cell in col:
			cell.font = Font(name='Calibri', size=11)


"""
function: infer data types for all columns for the input table
parameters:
_data: pandas DataFrame
	data table to infer
fname: string
	the output file name
output_root: string, default=''
	the root directory for the output file
sample_size: int or float(<= 1.0), default=1.0
	int: number of sample rows to infer the data type (useful for large tables)
	float: sample size in percentage
type_threshold: float(<= 1.0), default=0.5
	threshold for inferring data type
n_jobs: int, default=1
	the number of jobs to run in parallel
base_schema: pandas DataFrame, default=None
	data schema to base on
base_schema_feature_colname: string
	feature_colname in base schema
base_schema_dtype_colname: string
	dtype_colname in base schema
"""
def infer_schema(_data, fname, output_root='', sample_size=1.0, type_threshold=0.5, n_jobs=1, 
	base_schema=None, base_schema_feature_colname='column', base_schema_dtype_colname='type'):
	# check _data
	if type(_data) != pd.core.frame.DataFrame:
		raise ValueError('_data: only accept pandas DataFrame')

	# check sample_size
	if (type(sample_size) != int) and (type(sample_size) != float):
		raise ValueError('sample_size: only accept integer or float value')
	if sample_size > 1:
		if int(sample_size) != sample_size:
			raise ValueError('sample_size: only accept integer when it is > 1.0')
		if sample_size > _data.shape[0]:
			raise ValueError('sample_size: should be smaller or equal to len(_data)')
	else:
		if sample_size <= 0:
			raise ValueError('sample_size: should be larger than 0')

	# check fname
	if type(fname) != str:
		raise ValueError('fname: only accept string')

	# check output_root
	if output_root != '':
		if type(output_root) != str:
			raise ValueError('output_root: only accept string')
		if not os.path.isdir(output_root):
			raise ValueError('output_root: root not exists')

	# check type_threshold
	if type(type_threshold) != float:
		raise ValueError('type_threshold: only accept float value')
	if (type_threshold <= 0) or (type_threshold > 1):
		raise ValueError('type_threshold: should be in (0, 1]') 

	# check n_jobs
	if type(n_jobs) != int:
		raise ValueError('n_jobs: only accept integer value') 

	# check base_schema
	if base_schema is not None:
		if type(base_schema) != pd.core.frame.DataFrame:
			raise ValueError('base_schema: only accept pandas DataFrame')
		if type(base_schema_feature_colname) != str:
			raise ValueError('base_schema_feature_colname: only accept string value')
		if not base_schema_feature_colname in base_schema.columns.values:
			raise ValueError('base_schema_feature_colname: column not in base schema')
		if type(base_schema_dtype_colname) != str:
			raise ValueError('base_schema_dtype_colname: only accept string value')
		if not base_schema_dtype_colname in base_schema.columns.values:
			raise ValueError('base_schema_dtype_colname: column not in base schema')  

	# copy raw data table
	data = _data.copy()

	# open a new workbook to store all result
	wb = openpyxl.Workbook()
	ws = wb['Sheet']
	ws.title = 'schema'

	# calculate sample size
	if sample_size <= 1.0:
		sample_size = int(data.shape[0] * sample_size)

	# dictionary to store dropna sample data values
	data_dropna_sample_values = {}
	for col in data.columns.values:
		if len(data[col].dropna()) <= sample_size:
			data_dropna_sample_values[col] = data[col].dropna().values
		else:
			data_dropna_sample_values[col] = data[col].dropna().sample(sample_size).values
	
	# use data_dropna_sample_values to infer data type for each column
	type_infos = Parallel(n_jobs=n_jobs)(delayed(_infer_dtype)(data_dropna_sample_values[col], col, type_threshold) 
		for col in data.columns.values)
	type_infos_df = pd.DataFrame(type_infos)[['column', 'type']]

	# dtype mapping for basic stat calculation
	data_types = {}
	for col in data.columns.values:
		data_types[col] = type_infos_df.loc[type_infos_df['column']==col, 'type'].values[0]
	
	# get basic statistic information for all columns
	stat_infos = Parallel(n_jobs=n_jobs)(delayed(_cal_column_stat)
		(data_dropna_sample_values[col], col, data_types[col]) for col in data.columns.values)
	stat_infos_df = pd.DataFrame(stat_infos)
	
	# merge dtype infomation with stat information
	full_infos_df = type_infos_df.merge(stat_infos_df, on='column', how='left')
	full_infos_df = full_infos_df[['column', 'type', 'sample_value', 'sample_num_uni', 'sample_min', 'sample_median', 'sample_max', 'sample_std']]

	# if base_schema is provided, we can compare with base schema
	if base_schema is not None:
		base_schema = base_schema.rename(columns={base_schema_feature_colname: 'base_column', 
			base_schema_dtype_colname: 'base_type'})[['base_column', 'base_type']]
		full_infos_df = full_infos_df.merge(base_schema, left_on='column', right_on='base_column', how='outer')

		# compare with the base schema
		full_infos_df['base_column'] = full_infos_df['base_column'].apply(lambda x : 'column not in base table' if pd.isnull(x) else x)
		full_infos_df['column'] = full_infos_df['column'].apply(lambda x : 'column not in current table' if pd.isnull(x) else x)

		# reorder the column
		full_infos_df = full_infos_df[['column', 'base_column', 'type', 'base_type', 
		'sample_value', 'sample_num_uni', 'sample_min', 'sample_median', 'sample_max', 'sample_std']]


	# add data validation for type column
	val_type = DataValidation(type="list", formula1='"key,numeric,str,date"', allow_blank=False)
	ws.add_data_validation(val_type)

	# get col_name, excel column mapping
	column_mapping = {}
	for i, col in enumerate(full_infos_df.columns):
		column_mapping[col] = xlsxwriter.utility.xl_col_to_name(i)

	# write everything into the worksheet
	for r_idx, r in enumerate(dataframe_to_rows(full_infos_df, index=False, header=True)):
		ws.append(r)
		if r_idx != 0:
			val_type.add(ws['%s%d' %(column_mapping['type'], ws.max_row)])

		# bold the column names
		if r_idx == 0:
			for col in ws.iter_cols(max_col=ws.max_column, min_row=ws.max_row, max_row=ws.max_row):
				for cell in col:
					cell.font = Font(bold=True)

	# add conditional formating
	red_fill = PatternFill(bgColor="FFC7CE")
	red_font = Font(color="9C0006")
	green_fill = PatternFill(bgColor="C6EFCE")
	green_font = Font(color="006100")
	yellow_fill = PatternFill(bgColor="ffeb9c")
	yellow_font = Font(color="9c6500")

	# red highlight if there is any inconsistent between base and the target
	if base_schema is not None:
		col1 = column_mapping['column']
		col2 = column_mapping['base_column']
		ws.conditional_formatting.add('%s2:%s%d' %(col1, col1, ws.max_row), 
									  FormulaRule(formula=['%s2<>%s2' %(col1, col2)], stopIfTrue=True, fill=red_fill, font=red_font))
		
		ws.conditional_formatting.add('%s2:%s%d' %(col2, col2, ws.max_row), 
									  FormulaRule(formula=['%s2<>%s2' %(col1, col2)], stopIfTrue=True, fill=red_fill, font=red_font))
									  

		col1 = column_mapping['type']
		col2 = column_mapping['base_type']
		ws.conditional_formatting.add('%s2:%s%d' %(col1, col1, ws.max_row), 
									  FormulaRule(formula=['%s2<>%s2' %(col1, col2)], stopIfTrue=True, fill=red_fill, font=red_font))
		
		ws.conditional_formatting.add('%s2:%s%d' %(col2, col2, ws.max_row), 
									  FormulaRule(formula=['%s2<>%s2' %(col1, col2)], stopIfTrue=True, fill=red_fill, font=red_font))
									  

	# yellow hightlight column type (which need to be modified)
	ws['%s1' %(column_mapping['type'])].style = 'Neutral'

	# green highlight for the mkey type and red highlight for the error type
	type_cols = [column_mapping['type']]
	if 'base_type' in column_mapping.keys():
		type_cols.append(column_mapping['base_type'])

	for col in type_cols:
		ws.conditional_formatting.add('%s2:%s%d' %(col, col, ws.max_row), 
									  FormulaRule(formula=['%s2="error"' %(col)], stopIfTrue=True, fill=red_fill, font=red_font))
		ws.conditional_formatting.add('%s2:%s%d' %(col, col, ws.max_row), 
									  FormulaRule(formula=['%s2="key"' %(col)], stopIfTrue=True, fill=green_fill, font=green_font))

	# red highlight for sample_num_uni = 0 or 1, only one unique value
	ws.conditional_formatting.add('%s2:%s%d' %(column_mapping['sample_num_uni'], column_mapping['sample_num_uni'], ws.max_row), 
								  FormulaRule(formula=['%s2=0' %(column_mapping['sample_num_uni'])], stopIfTrue=True, fill=red_fill, font=red_font))
	ws.conditional_formatting.add('%s2:%s%d' %(column_mapping['sample_num_uni'], column_mapping['sample_num_uni'], ws.max_row), 
								  FormulaRule(formula=['%s2=1' %(column_mapping['sample_num_uni'])], stopIfTrue=True, fill=red_fill, font=red_font))

	# adjust the column format for the worksheet
	_adjust_column(ws)

	wb.save(filename = os.path.join(output_root, 'data_schema_%s.xlsx' %(fname)))

